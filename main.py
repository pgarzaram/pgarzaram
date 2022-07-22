# Python Program Update Inventory Stocks
# import openpyxl module
import openpyxl
import actionfunctions

# Location of a file
path = "C:/Users/Pedro/Desktop/testing/Inventory_Management/Inventory_Sheet.xlsx"

#Activate Sheet
wb = openpyxl.load_workbook(path)
ws = wb.active
# Display where in workbook you are at
print(f"You are currently on Sheet: {ws.title} ")

StockItems_PartNumber = ws.cell(row = 1, column = 1)
StockItems_PartDescription = ws.cell(row = 1, column = 2)
StockItems_ItemsInStock = ws.cell(row = 1, column = 3)

Column_Titles = [StockItems_PartNumber.value, StockItems_PartDescription.value, StockItems_ItemsInStock.value]
print(Column_Titles)


max_row = ws.max_row
max_column = ws.max_column
print(max_row)
print(max_column)

for i in range(1, max_row+1):
    column_1 = ws.cell(row = i, column = 1)
    column_2 = ws.cell(row = i, column = 2)
    column_3 = ws.cell(row = i, column = 3)
    print(["Part Number: " + str(column_1.value), "Part Description: " + str(column_2.value), "In Stock: " + str(column_3.value)])

ActionOut = input("Would you like to check items out of inventory? Y/N\n")
ActionIn = input("Would you like to check add items to inventory? Y/N\n")


if ActionOut == "Y" and ActionIn =="N":
    actionfunctions.CheckOut(max_row,ws)
elif ActionOut == "N" and ActionIn == "Y":
    actionfunctions.CheckIn(max_row,ws)
elif ActionOut == "Y" and ActionIn == "Y":
    actionfunctions.CheckOut(max_row, ws)
    actionfunctions.CheckIn(max_row, ws)
else :
    print("The session has ended.")




wb.save(path)